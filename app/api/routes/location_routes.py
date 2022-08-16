from openpyxl import load_workbook
from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from app.api.dependencies.db import get_db
import os
from app.models.collection_centres import CollectionCentres
from app.models.local_government import LocalGovernment
from app.schema.visits_schemas import LocalGovt, CollectionCent


router = APIRouter()


@router.post("/create_collection_Centres")
def create_collection_centres(db: Session = Depends(get_db)):
    location = os.getcwd()
    wb = load_workbook(f'{location}/excel_files/CollectionCentres.xlsx')
    ws = wb.active
    collection_centres = ws["D"]
    for i in range(2, len(collection_centres)+1):
        name = ws[f"D{i}"].value
        code = ws[f"B{i}"].value
        localGovtCode = ws[f"C{i}"].value
        db_collection_centres = CollectionCentres(
            name=name, code=code, localGovtCode=localGovtCode)
        db.add(db_collection_centres)
        db.commit()
        db.refresh(db_collection_centres)

    return {"message": "Collection centres Created Successfully"}


@router.get("/get_all_collection_centres")
def get_all_collection_centres(db: Session = Depends(get_db)):
    collection_centres = db.query(CollectionCentres).all()

    return [
        CollectionCent(
            id=collection_centre.id,
            name=collection_centre.name,
            code=collection_centre.code,
            lgCode=collection_centre.localGovtCode)
        for collection_centre in collection_centres
    ]


@router.get("/get_collection_centre")
def get_collection_centre(code: str, db: Session = Depends(get_db)):
    collection_centre = db.query(CollectionCentres).filter(
        CollectionCentres.code == code).all()
    return collection_centre
    # return CollectionCent(
    #     id=collection_centre.id,
    #     name=collection_centre.name,
    #     code=collection_centre.code,
    #     lgCode=collection_centre.localGovtCode)


@router.get('/get_local_governments')
def get_all_local_governments(db: Session = Depends(get_db)):
    local_governments = db.query(LocalGovernment).all()
    return [
        LocalGovt(
            id=local_govt.id,
            name=local_govt.name,
            code=local_govt.code)
        for local_govt in local_governments
    ]


@router.post('create_localGovernments')
def create_local_government(db: Session = Depends(get_db)):
    location = os.getcwd()
    wb = load_workbook(f'{location}/excel_files/local_govts.xlsx')
    ws = wb.active
    local_government = ws['B']
    for i in range(2, len(local_government)+1):
        name = ws[f'B{i}'].value
        code = ws[f'D{i}'].value
        db_local_govt = LocalGovernment(name=name, code=code)
        db.add(db_local_govt)
        db.commit()
        db.refresh(db_local_govt)

    return {"message": "Local Governments Created Successfully"}
